"""
Excel Gantt для Балаково
"""

from datetime import datetime, date, timedelta
from typing import List, Dict, Optional
from pathlib import Path
import calendar
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from modules.balakovo_data import BalakovoData, BerthSchedule, BerthingSlot

logger = logging.getLogger(__name__)


# Цвета
COLORS = {
    'barge_sfo': '92D050',     # Зелёный - баржа SFO
    'barge_rpo': '00B050',     # Тёмно-зелёный - баржа RPO
    'dry_meal': 'FFC000',      # Оранжевый - сухогруз MEAL
    'dry_pellets': 'FF6600',   # Тёмно-оранжевый - PELLETS
    'waiting': 'FF0000',       # Красный - ожидание
    'restriction': 'A6A6A6',   # Серый - ограничение
}

FILLS = {k: PatternFill(start_color=v, end_color=v, fill_type='solid') 
         for k, v in COLORS.items()}

HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=9)
WEEKEND_FILL = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')


class BalakovoGanttExcel:
    """Генератор Excel Gantt для Балаково"""
    
    def __init__(self, data: BalakovoData, output_dir: str = "output/balakovo"):
        self.data = data
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_schedule(self) -> str:
        """Генерация расписания причалов"""
        logger.info("\n Генерация Excel Gantt...")
        
        wb = Workbook()
        
        # Лист 1: Gantt по дням
        self._create_gantt_sheet(wb)
        
        # Лист 2: Список постановок
        self._create_slots_sheet(wb)
        
        # Лист 3: Конфликты
        self._create_conflicts_sheet(wb)
        
        # Удаляем пустой лист
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        filepath = self.output_dir / "berth_schedule.xlsx"
        wb.save(filepath)
        
        logger.info(f"   {filepath}")
        return str(filepath)
    
    def _create_gantt_sheet(self, wb: Workbook):
        """Создание Gantt листа"""
        ws = wb.create_sheet("Gantt")
        
        all_dates = self._get_all_dates()
        if not all_dates:
            ws['A1'] = "Нет запланированных постановок"
            return
        
        start_date = min(all_dates)
        end_date = max(all_dates)
        
        # Setup header
        header_row = self._setup_gantt_header(ws, start_date, end_date)
        
        # Fill data
        next_row = self._fill_gantt_data(ws, header_row + 1, start_date, end_date)
        
        # Add legend
        self._add_gantt_legend(ws, next_row + 2)

    def _get_all_dates(self) -> List[date]:
        """Get all relevant dates from schedules."""
        all_dates = []
        for schedule in self.data.schedules.values():
            for slot in schedule.slots:
                all_dates.append(slot.berthing_start.date())
                all_dates.append(slot.departure.date())
        return all_dates

    def _setup_gantt_header(self, ws, start_date: date, end_date: date) -> int:
        """Setup Gantt chart header with months and days."""
        # Title
        ws['A1'] = f"РАСПИСАНИЕ ПРИЧАЛОВ БАЛАКОВО"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Период: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
        
        header_row = 4
        
        # Column A header
        ws.cell(row=header_row, column=1, value="Причал").fill = HEADER_FILL
        ws.cell(row=header_row, column=1).font = HEADER_FONT
        ws.column_dimensions['A'].width = 20
        
        # Days header
        current = start_date
        col = 2
        while current <= end_date:
            cell = ws.cell(row=header_row, column=col, value=current.day)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(col)].width = 4
            
            # Weekend highlight in header
            if current.weekday() >= 5:
                cell.fill = PatternFill(start_color='4A6FA5', end_color='4A6FA5', fill_type='solid')
            
            current += timedelta(days=1)
            col += 1
            
        # Months header
        month_row = header_row - 1
        current = start_date
        col = 2
        while current <= end_date:
            if current.day == 1 or current == start_date:
                month_name = current.strftime('%B %Y')
                ws.cell(row=month_row, column=col, value=month_name)
                ws.cell(row=month_row, column=col).font = Font(bold=True)
            current += timedelta(days=1)
            col += 1
            
        return header_row

    def _fill_gantt_data(self, ws, start_row: int, start_date: date, end_date: date) -> int:
        """Fill Gantt chart with berth schedules."""
        current_row = start_row
        
        for berth_id in sorted(self.data.schedules.keys()):
            schedule = self.data.schedules[berth_id]
            berth = self.data.berths[berth_id]
            
            # Berth name
            cell = ws.cell(row=current_row, column=1, value=berth.berth_name)
            cell.font = Font(bold=True)
            
            # Fill days
            current = start_date
            col = 2
            while current <= end_date:
                cell = ws.cell(row=current_row, column=col)
                
                # Weekend background
                if current.weekday() >= 5:
                    cell.fill = WEEKEND_FILL
                
                # Check restrictions and slots
                self._fill_day_cell(cell, current, berth_id, schedule)
                
                current += timedelta(days=1)
                col += 1
            
            current_row += 1
            
        return current_row

    def _fill_day_cell(self, cell, current_date: date, berth_id: str, schedule: BerthSchedule):
        """Fill a single cell in the Gantt chart."""
        # Restrictions
        if self.data.is_date_blocked(current_date, berth_id):
            cell.fill = FILLS['restriction']
            cell.value = ''
            cell.alignment = Alignment(horizontal='center')
            return

        # Slots
        for slot in schedule.slots:
            if slot.berthing_start.date() <= current_date <= slot.departure.date():
                # Determine color
                if slot.cargo_type in ['SFO', 'RPO', 'CSO']:
                    fill_key = 'barge_sfo' if slot.cargo_type == 'SFO' else 'barge_rpo'
                else:
                    fill_key = 'dry_meal' if slot.cargo_type == 'MEAL' else 'dry_pellets'
                
                cell.fill = FILLS[fill_key]
                
                # Label on first day
                if slot.berthing_start.date() == current_date:
                    cell.value = slot.vessel_name[:3]
                
                cell.font = Font(size=7, bold=True, color='FFFFFF')
                cell.alignment = Alignment(horizontal='center')
                break

    def _add_gantt_legend(self, ws, row: int):
        """Add legend to Gantt chart."""
        ws.cell(row=row, column=1, value="Легенда:").font = Font(bold=True)
        
        legend = [
            ('Баржа SFO → Olya', 'barge_sfo'),
            ('Баржа RPO → Olya', 'barge_rpo'),
            ('Сухогруз MEAL → Turkey', 'dry_meal'),
            ('Сухогруз PELLETS → Turkey', 'dry_pellets'),
            ('Ограничение', 'restriction'),
        ]
        
        for i, (label, color_key) in enumerate(legend):
            cell = ws.cell(row=row, column=2 + i * 2, value=label)
            cell.fill = FILLS[color_key]
            cell.font = Font(size=8, color='FFFFFF' if color_key != 'restriction' else '000000')
    
    def _create_slots_sheet(self, wb: Workbook):
        """Список постановок"""
        ws = wb.create_sheet("Постановки")
        
        headers = [
            'Слот', 'Причал', 'Судно', 'Груз', 'Кол-во (MT)', 'Направление',
            'Прибытие', 'Швартовка', 'Начало погрузки', 'Конец погрузки',
            'Отход', 'Ожидание (ч)', 'Погрузка (ч)', 'Статус'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        
        row = 2
        for schedule in self.data.schedules.values():
            for slot in sorted(schedule.slots, key=lambda s: s.berthing_start):
                berth = self.data.berths[slot.berth_id]
                
                ws.cell(row=row, column=1, value=slot.slot_id)
                ws.cell(row=row, column=2, value=berth.berth_name)
                ws.cell(row=row, column=3, value=slot.vessel_name)
                ws.cell(row=row, column=4, value=slot.cargo_type)
                ws.cell(row=row, column=5, value=slot.qty_mt)
                ws.cell(row=row, column=6, value=slot.destination)
                ws.cell(row=row, column=7, value=slot.eta.strftime('%d.%m %H:%M'))
                ws.cell(row=row, column=8, value=slot.berthing_start.strftime('%d.%m %H:%M'))
                ws.cell(row=row, column=9, value=slot.loading_start.strftime('%d.%m %H:%M'))
                ws.cell(row=row, column=10, value=slot.loading_end.strftime('%d.%m %H:%M'))
                ws.cell(row=row, column=11, value=slot.departure.strftime('%d.%m %H:%M'))
                ws.cell(row=row, column=12, value=round(slot.waiting_hours, 1))
                ws.cell(row=row, column=13, value=round(slot.loading_hours, 1))
                ws.cell(row=row, column=14, value=slot.status.value)
                
                row += 1
        
        # Автоширина
        for col in range(1, 15):
            ws.column_dimensions[get_column_letter(col)].width = 14
    
    def _create_conflicts_sheet(self, wb: Workbook):
        """Список конфликтов"""
        ws = wb.create_sheet("Конфликты")
        
        headers = ['Груз', 'Тип', 'Описание', 'Серьёзность']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        
        row = 2
        for conflict in self.data.conflicts:
            ws.cell(row=row, column=1, value=conflict['cargo_id'])
            ws.cell(row=row, column=2, value=conflict['type'])
            ws.cell(row=row, column=3, value=conflict['description'])
            ws.cell(row=row, column=4, value=conflict['severity'])
            
            # Подсветка ошибок
            if conflict['severity'] == 'error':
                for col in range(1, 5):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'
                    )
            
            row += 1
        
        if not self.data.conflicts:
            ws.cell(row=2, column=1, value="Конфликтов нет ")
            ws.cell(row=2, column=1).font = Font(color='006600')
        
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 20
