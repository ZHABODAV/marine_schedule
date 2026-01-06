"""
Balakovo Report Module
Generates Balakovo-specific reports for meal and oil operations.
"""
from __future__ import annotations

import pandas as pd
from datetime import datetime
from typing import Optional, Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class BalakovoReportGenerator:
    """Generates specialized reports for Balakovo operations."""
    
    def __init__(self, voyage_data: pd.DataFrame):
        """
        Initialize with voyage data.
        
        Args:
            voyage_data: DataFrame with voyage information
        """
        self.voyage_data = voyage_data.copy()
        self._prepare_data()
    
    def _prepare_data(self) -> None:
        """Prepare and validate data for Balakovo reporting."""
        if 'start_time' in self.voyage_data.columns:
            self.voyage_data['start_time'] = pd.to_datetime(self.voyage_data['start_time'])
        if 'end_time' in self.voyage_data.columns:
            self.voyage_data['end_time'] = pd.to_datetime(self.voyage_data['end_time'])
    
    def filter_balakovo_operations(self, port_name: str = 'Balakovo') -> pd.DataFrame:
        """
        Filter for Balakovo-specific operations.
        
        Args:
            port_name: Name of Balakovo port
        
        Returns:
            Filtered DataFrame
        """
        df = self.voyage_data
        
        # Check for port columns and filter
        mask = pd.Series([False] * len(df), index=df.index)
        if 'start_port' in df.columns:
            mask |= df['start_port'].astype(str).str.contains(port_name, case=False, na=False)
        if 'end_port' in df.columns:
            mask |= df['end_port'].astype(str).str.contains(port_name, case=False, na=False)
        
        balakovo_ops = df[mask]
        
        return balakovo_ops
    
    def generate_meal_operations_report(self, port_name: str = 'Balakovo') -> pd.DataFrame:
        """
        Generate report for meal operations at Balakovo.
        
        Args:
            port_name: Port name
        
        Returns:
            DataFrame with meal operations report
        """
        balakovo_ops = self.filter_balakovo_operations(port_name)
        
        if balakovo_ops.empty:
            return pd.DataFrame()
        
        # Identify meal operations (typically loading operations)
        mask = pd.Series([False] * len(balakovo_ops), index=balakovo_ops.index)
        if 'leg_type' in balakovo_ops.columns:
            mask |= balakovo_ops['leg_type'].astype(str).str.contains('load', case=False, na=False)
        if 'cargo_type' in balakovo_ops.columns:
            mask |= balakovo_ops['cargo_type'].astype(str).str.contains('meal', case=False, na=False)
        
        meal_ops = balakovo_ops[mask].copy()
        
        # Calculate metrics
        if not meal_ops.empty:
            meal_ops['operation_date'] = meal_ops['start_time'].dt.date  # type: ignore[attr-defined]
            meal_ops['duration_days'] = (
                (meal_ops['end_time'] - meal_ops['start_time']).dt.total_seconds() / 86400  # type: ignore[attr-defined]
            )
            
            # Add cargo quantity if available
            if 'cargo_quantity' not in meal_ops.columns:
                meal_ops['cargo_quantity'] = 0
            
            # Calculate loading rate
            meal_ops['loading_rate_per_day'] = meal_ops.apply(
                lambda x: x.get('cargo_quantity', 0) / x['duration_days'] if x['duration_days'] > 0 else 0,
                axis=1
            )
        
        return meal_ops
    
    def generate_oil_operations_report(self, port_name: str = 'Balakovo') -> pd.DataFrame:
        """
        Generate report for oil operations at Balakovo.
        
        Args:
            port_name: Port name
        
        Returns:
            DataFrame with oil operations report
        """
        balakovo_ops = self.filter_balakovo_operations(port_name)
        
        if balakovo_ops.empty:
            return pd.DataFrame()
        
        # Identify oil operations
        mask = pd.Series([False] * len(balakovo_ops), index=balakovo_ops.index)
        if 'cargo_type' in balakovo_ops.columns:
            mask = balakovo_ops['cargo_type'].astype(str).str.contains('oil', case=False, na=False)
        
        oil_ops = balakovo_ops[mask].copy()
        
        if not oil_ops.empty:
            oil_ops['operation_date'] = oil_ops['start_time'].dt.date  # type: ignore[attr-defined]
            oil_ops['duration_days'] = (
                (oil_ops['end_time'] - oil_ops['start_time']).dt.total_seconds() / 86400  # type: ignore[attr-defined]
            )
            
            # Add cargo quantity if available
            if 'cargo_quantity' not in oil_ops.columns:
                oil_ops['cargo_quantity'] = 0
            
            # Calculate loading/unloading rate
            oil_ops['operation_rate_per_day'] = oil_ops.apply(
                lambda x: x.get('cargo_quantity', 0) / x['duration_days'] if x['duration_days'] > 0 else 0,
                axis=1
            )
        
        return oil_ops
    
    def generate_monthly_summary(self, port_name: str = 'Balakovo', 
                                year: Optional[int] = None,
                                month: Optional[int] = None) -> pd.DataFrame:
        """
        Generate monthly summary for Balakovo operations.
        
        Args:
            port_name: Port name
            year: Optional year filter
            month: Optional month filter
        
        Returns:
            DataFrame with monthly summary
        """
        balakovo_ops = self.filter_balakovo_operations(port_name)
        
        if balakovo_ops.empty:
            return pd.DataFrame()
        
        # Add month/year columns
        balakovo_ops['year'] = balakovo_ops['start_time'].dt.year  # type: ignore[attr-defined]
        balakovo_ops['month'] = balakovo_ops['start_time'].dt.month  # type: ignore[attr-defined]
        
        # Filter if specified
        if year:
            balakovo_ops = balakovo_ops[balakovo_ops['year'] == year]
        if month:
            balakovo_ops = balakovo_ops[balakovo_ops['month'] == month]
        
        # Group by month and cargo type
        summary = balakovo_ops.groupby(['year', 'month', 'cargo_type']).agg({
            'asset': 'count',  # Number of operations
            'cargo_quantity': 'sum',  # Total cargo
            'duration_hours': 'sum'  # Total hours
        }).reset_index()
        
        summary.columns = ['year', 'month', 'cargo_type', 'total_operations', 
                          'total_cargo', 'total_hours']
        
        return summary
    
    def export_balakovo_report(self, output_file: str, port_name: str = 'Balakovo') -> None:
        """
        Export comprehensive Balakovo report to Excel.
        
        Args:
            output_file: Output Excel file path
            port_name: Port name
        """
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Overview sheet
            balakovo_all = self.filter_balakovo_operations(port_name)
            if not balakovo_all.empty:
                balakovo_all.to_excel(writer, sheet_name='All Operations', index=False)
            
            # Meal operations
            meal_ops = self.generate_meal_operations_report(port_name)
            if not meal_ops.empty:
                meal_ops.to_excel(writer, sheet_name='Meal Operations', index=False)
            
            # Oil operations
            oil_ops = self.generate_oil_operations_report(port_name)
            if not oil_ops.empty:
                oil_ops.to_excel(writer, sheet_name='Oil Operations', index=False)
            
            # Monthly summary
            monthly = self.generate_monthly_summary(port_name)
            if not monthly.empty:
                monthly.to_excel(writer, sheet_name='Monthly Summary', index=False)
            
            # Format worksheets
            self._format_workbook(writer.book)
    
    def _format_workbook(self, workbook: Workbook) -> None:
        """
        Apply formatting to workbook.
        
        Args:
            workbook: openpyxl Workbook object
        """
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for sheet in workbook.worksheets:
            # Format headers
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-size columns
            for column in sheet.columns:
                max_length = 0
                col_idx = column[0].column
                if col_idx is None:
                    continue
                column_letter = get_column_letter(col_idx)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (AttributeError, TypeError):
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
    
    def get_operation_statistics(self, port_name: str = 'Balakovo') -> Dict:
        """
        Get statistical summary for Balakovo operations.
        
        Args:
            port_name: Port name
        
        Returns:
            Dictionary with statistics
        """
        balakovo_ops = self.filter_balakovo_operations(port_name)
        
        if balakovo_ops.empty:
            return {}
        
        stats = {
            'total_operations': len(balakovo_ops),
            'unique_assets': balakovo_ops['asset'].nunique() if 'asset' in balakovo_ops.columns else 0,
            'total_cargo': balakovo_ops['cargo_quantity'].sum() if 'cargo_quantity' in balakovo_ops.columns else 0,
            'avg_operation_hours': balakovo_ops['duration_hours'].mean() if 'duration_hours' in balakovo_ops.columns else 0,
            'date_range': {
                'start': balakovo_ops['start_time'].min() if 'start_time' in balakovo_ops.columns else None,
                'end': balakovo_ops['end_time'].max() if 'end_time' in balakovo_ops.columns else None
            }
        }
        
        # Cargo type breakdown
        if 'cargo_type' in balakovo_ops.columns:
            cargo_breakdown = balakovo_ops.groupby('cargo_type').size().to_dict()
            stats['cargo_type_breakdown'] = cargo_breakdown
        
        return stats


def generate_balakovo_report(voyage_data: pd.DataFrame, 
                             output_file: str,
                             port_name: str = 'Balakovo') -> BalakovoReportGenerator:
    """
    Generate Balakovo-specific reports.
    
    Args:
        voyage_data: DataFrame with voyage data
        output_file: Output Excel file path
        port_name: Port name for filtering
    
    Returns:
        BalakovoReportGenerator instance
    """
    generator = BalakovoReportGenerator(voyage_data)
    generator.export_balakovo_report(output_file, port_name)
    
    return generator
