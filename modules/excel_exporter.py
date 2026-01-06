"""
Экспорт расписания в Excel
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import logging

from modules.data_loader import AllData, Voyage, VoyageLeg
from modules.olya_coordinator import OlyaSchedule, StowageMatch

logger = logging.getLogger(__name__)


# Стили
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
OK_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
CRITICAL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


class ExcelExporter:
    """Экспорт в Excel"""
    
    def __init__(self, data: AllData, output_dir: str = "output"):
        self.data = data
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def export_vessel_schedule(self, filename: str = "vessel_schedule.xlsx") -> str:
        """Экспорт расписания судов"""
        logger.info("Экспорт расписания судов в Excel...")
        
        # Готовим данные
        rows = []
        
        for voyage in sorted(self.data.voyages.values(), key=lambda v: (v.vessel_id, v.start_time or datetime.min)):
            vessel = self.data.vessels.get(voyage.vessel_id)
            
            for leg in voyage.legs:
                rows.append({
                    'Судно': voyage.vessel_name,
                    'Класс': vessel.vessel_class if vessel else '',
                    'Рейс': voyage.voyage_id,
                    'Leg': leg.leg_seq,
                    'Операция': leg.op_detail,
                    'Тип': leg.leg_type,
                    'Порт Начало': leg.port_start_id,
                    'Порт Конец': leg.port_end_id,
                    'Груз': leg.cargo_id or '',
                    'Кол-во MT': leg.qty_mt,
                    'Начало': leg.start_time.strftime('%Y-%m-%d %H:%M') if leg.start_time else '',
                    'Окончание': leg.end_time.strftime('%Y-%m-%d %H:%M') if leg.end_time else '',
                    'Длительность (дни)': round(leg.duration_days, 1),
                    'Статус': leg.status,
                    'Примечания': leg.remarks
                })
        
        df = pd.DataFrame(rows)
        
        # Создаем Excel
        filepath = self.output_dir / filename
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Schedule', index=False)
            
            # Форматирование
            wb = writer.book
            ws = writer.sheets['Schedule']
            
            # Заголовки
            for cell in ws[1]:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal='center')
                cell.border = THIN_BORDER
            
            # Ширина колонок
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['K'].width = 16
            ws.column_dimensions['L'].width = 16
            ws.column_dimensions['O'].width = 30
            
            # Закрепляем заголовок
            ws.freeze_panes = 'A2'
        
        logger.info(f"   Сохранено: {filepath}")
        return str(filepath)
    
    def export_voyage_summary(self, filename: str = "voyage_summary.xlsx") -> str:
        """Экспорт сводки по рейсам"""
        logger.info("Экспорт сводки по рейсам...")
        
        rows = []
        
        for voyage in sorted(self.data.voyages.values(), key=lambda v: v.start_time or datetime.min):
            vessel = self.data.vessels.get(voyage.vessel_id)
            
            rows.append({
                'Рейс ID': voyage.voyage_id,
                'Судно': voyage.vessel_name,
                'Класс': vessel.vessel_class if vessel else '',
                'Тип контракта': vessel.contract_type if vessel else '',
                'Порт загрузки': voyage.load_port,
                'Порт выгрузки': voyage.discharge_port,
                'Груз': voyage.cargo_id,
                'Кол-во MT': voyage.qty_mt,
                'Начало': voyage.start_time.strftime('%Y-%m-%d') if voyage.start_time else '',
                'Окончание': voyage.end_time.strftime('%Y-%m-%d') if voyage.end_time else '',
                'Этапов': len(voyage.legs),
                'Статус': voyage.status
            })
        
        df = pd.DataFrame(rows)
        
        filepath = self.output_dir / filename
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Voyages', index=False)
            
            wb = writer.book
            ws = writer.sheets['Voyages']
            
            for cell in ws[1]:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal='center')
            
            ws.freeze_panes = 'A2'
        
        logger.info(f"   Сохранено: {filepath}")
        return str(filepath)
    
    def export_olya_schedule(
        self,
        schedule: OlyaSchedule,
        filename: str = "olya_coordination.xlsx"
    ) -> str:
        """Экспорт расписания стыковки в Olya"""
        logger.info("Экспорт расписания Olya...")
        
        wb = Workbook()
        
        # === Лист 1: Стыковки ===
        ws_matches = wb.active
        ws_matches.title = "Стыковки"
        
        headers = ['Статус', 'Баржа', 'Судно', 'Груз', 'Кол-во MT', 
                   'Разгрузка баржи', 'Загрузка судна', 'Буфер (часы)', 'Примечания']
        
        for col, header in enumerate(headers, 1):
            cell = ws_matches.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
        
        row = 2
        for match in schedule.matches:
            status_map = {'OK': '', 'WARNING': '', 'CRITICAL': '', 'OVERLAP': ''}
            fill_map = {'OK': OK_FILL, 'WARNING': WARNING_FILL, 'CRITICAL': CRITICAL_FILL, 'OVERLAP': CRITICAL_FILL}
            
            ws_matches.cell(row=row, column=1, value=status_map.get(match.status, '?'))
            ws_matches.cell(row=row, column=2, value=match.barge_op.vessel_name)
            ws_matches.cell(row=row, column=3, value=match.vessel_op.vessel_name)
            ws_matches.cell(row=row, column=4, value=match.vessel_op.cargo_id or '')
            ws_matches.cell(row=row, column=5, value=match.vessel_op.qty_mt)
            ws_matches.cell(row=row, column=6, value=match.barge_op.end_time.strftime('%d.%m %H:%M'))
            ws_matches.cell(row=row, column=7, value=match.vessel_op.start_time.strftime('%d.%m %H:%M'))
            ws_matches.cell(row=row, column=8, value=round(match.gap_hours, 1))
            ws_matches.cell(row=row, column=9, value=match.notes)
            
            # Цвет строки
            fill = fill_map.get(match.status)
            if fill:
                for col in range(1, 10):
                    ws_matches.cell(row=row, column=col).fill = fill
                    ws_matches.cell(row=row, column=col).border = THIN_BORDER
            
            row += 1
        
        # === Лист 2: Операции барж ===
        ws_barges = wb.create_sheet("Баржи")
        
        headers = ['Судно', 'Класс', 'Рейс', 'Груз', 'Кол-во MT', 'Начало', 'Окончание', 'Длит. (ч)']
        for col, header in enumerate(headers, 1):
            cell = ws_barges.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        
        row = 2
        for op in schedule.barge_ops:
            ws_barges.cell(row=row, column=1, value=op.vessel_name)
            ws_barges.cell(row=row, column=2, value=op.vessel_class)
            ws_barges.cell(row=row, column=3, value=op.voyage_id)
            ws_barges.cell(row=row, column=4, value=op.cargo_id or '')
            ws_barges.cell(row=row, column=5, value=op.qty_mt)
            ws_barges.cell(row=row, column=6, value=op.start_time.strftime('%d.%m %H:%M'))
            ws_barges.cell(row=row, column=7, value=op.end_time.strftime('%d.%m %H:%M'))
            ws_barges.cell(row=row, column=8, value=round(op.duration_hours, 1))
            row += 1
        
        # === Лист 3: Операции судов ===
        ws_vessels = wb.create_sheet("Суда")
        
        for col, header in enumerate(headers, 1):
            cell = ws_vessels.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        
        row = 2
        for op in schedule.vessel_ops:
            ws_vessels.cell(row=row, column=1, value=op.vessel_name)
            ws_vessels.cell(row=row, column=2, value=op.vessel_class)
            ws_vessels.cell(row=row, column=3, value=op.voyage_id)
            ws_vessels.cell(row=row, column=4, value=op.cargo_id or '')
            ws_vessels.cell(row=row, column=5, value=op.qty_mt)
            ws_vessels.cell(row=row, column=6, value=op.start_time.strftime('%d.%m %H:%M'))
            ws_vessels.cell(row=row, column=7, value=op.end_time.strftime('%d.%m %H:%M'))
            ws_vessels.cell(row=row, column=8, value=round(op.duration_hours, 1))
            row += 1
        
        # Сохраняем
        filepath = self.output_dir / filename
        wb.save(filepath)
        
        logger.info(f"   Сохранено: {filepath}")
        return str(filepath)