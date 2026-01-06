"""
Excel Gantt Chart Module
Creates Gantt charts in Excel for assets and berths.
"""
from __future__ import annotations

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from typing import Optional, List, Dict


class ExcelGanttChart:
    """Creates Gantt charts in Excel format."""
    
    def __init__(self):
        self.workbook = Workbook()
        self.colors = {
            'sailing': 'ADD8E6',      # Light blue
            'loading': '90EE90',       # Light green
            'unloading': 'FFB6C1',     # Light pink
            'waiting': 'FFFFE0',       # Light yellow
            'maintenance': 'FFD700',   # Gold
            'default': 'D3D3D3'        # Light gray
        }
    
    def create_asset_gantt(self, voyage_data: pd.DataFrame, 
                          start_date: Optional[datetime] = None,
                          end_date: Optional[datetime] = None) -> None:
        """
        Create Gantt chart for assets.
        
        Args:
            voyage_data: DataFrame with voyage information
            start_date: Optional start date for the chart
            end_date: Optional end date for the chart
        """
        ws = self.workbook.active
        ws.title = "Asset Gantt"
        
        # Prepare data
        df = voyage_data.copy()
        df['start_time'] = pd.to_datetime(df['start_time'])
        df['end_time'] = pd.to_datetime(df['end_time'])
        
        # Determine date range
        if start_date is None:
            start_date = df['start_time'].min()
        if end_date is None:
            end_date = df['end_time'].max()
        
        # Create date columns (daily)
        date_range = pd.date_range(start=start_date, end=end_date, freq='D')
        
        # Headers
        ws['A1'] = 'Asset'
        ws['B1'] = 'Activity'
        for idx, date in enumerate(date_range, start=3):
            col = get_column_letter(idx)
            ws[f'{col}1'] = date.strftime('%Y-%m-%d')
            ws[f'{col}1'].alignment = Alignment(horizontal='center', textRotation=90)
        
        # Group by asset
        row = 2
        for asset in sorted(df['asset'].unique()):
            asset_data = df[df['asset'] == asset].sort_values('start_time')
            
            for _, leg in asset_data.iterrows():
                ws[f'A{row}'] = asset
                ws[f'B{row}'] = f"{leg['start_port']} → {leg['end_port']}"
                
                # Fill cells for duration
                leg_start = leg['start_time']
                leg_end = leg['end_time']
                
                for idx, date in enumerate(date_range, start=3):
                    if leg_start.date() <= date.date() <= leg_end.date():
                        col = get_column_letter(idx)
                        cell = ws[f'{col}{row}']
                        
                        # Color based on leg type
                        leg_type = leg.get('leg_type', 'default')
                        color = self.colors.get(leg_type, self.colors['default'])
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        cell.value = '█'
                        cell.alignment = Alignment(horizontal='center')
                
                row += 1
        
        # Format headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
    
    def create_berth_gantt(self, voyage_data: pd.DataFrame,
                          start_date: Optional[datetime] = None,
                          end_date: Optional[datetime] = None) -> None:
        """
        Create Gantt chart for berth utilization.
        
        Args:
            voyage_data: DataFrame with voyage/berth information
            start_date: Optional start date for the chart
            end_date: Optional end date for the chart
        """
        ws = self.workbook.create_sheet("Berth Gantt")
        
        # Prepare data - focus on port activities
        df = voyage_data.copy()
        df['start_time'] = pd.to_datetime(df['start_time'])
        df['end_time'] = pd.to_datetime(df['end_time'])
        
        # Determine date range
        if start_date is None:
            start_date = df['start_time'].min()
        if end_date is None:
            end_date = df['end_time'].max()
        
        # Create date columns
        date_range = pd.date_range(start=start_date, end=end_date, freq='D')
        
        # Headers
        ws['A1'] = 'Port/Berth'
        ws['B1'] = 'Asset'
        for idx, date in enumerate(date_range, start=3):
            col = get_column_letter(idx)
            ws[f'{col}1'] = date.strftime('%Y-%m-%d')
            ws[f'{col}1'].alignment = Alignment(horizontal='center', textRotation=90)
        
        # Get unique ports (both start and end)
        ports = set()
        if 'start_port' in df.columns:
            ports.update(df['start_port'].unique())
        if 'end_port' in df.columns:
            ports.update(df['end_port'].unique())
        
        row = 2
        for port in sorted(ports):
            # Activities at this port
            port_activities = df[(df['start_port'] == port) | (df['end_port'] == port)]
            
            for _, activity in port_activities.iterrows():
                ws[f'A{row}'] = port
                ws[f'B{row}'] = activity['asset']
                
                # Fill cells for activity duration
                activity_start = activity['start_time']
                activity_end = activity['end_time']
                
                for idx, date in enumerate(date_range, start=3):
                    if activity_start.date() <= date.date() <= activity_end.date():
                        col = get_column_letter(idx)
                        cell = ws[f'{col}{row}']
                        
                        # Determine color
                        if activity.get('end_port') == port:
                            color = self.colors.get('loading', self.colors['default'])
                        else:
                            color = self.colors.get('unloading', self.colors['default'])
                        
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        cell.value = '█'
                        cell.alignment = Alignment(horizontal='center')
                
                row += 1
        
        # Format headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
    
    def save(self, filename: str) -> None:
        """Save the workbook to file."""
        self.workbook.save(filename)


def create_gantt_charts(voyage_data: pd.DataFrame, output_file: str,
                       include_assets: bool = True, include_berths: bool = True) -> None:
    """
    Create Gantt charts and save to Excel.
    
    Args:
        voyage_data: DataFrame with voyage information
        output_file: Output Excel file path
        include_assets: Whether to include asset Gantt chart
        include_berths: Whether to include berth Gantt chart
    """
    gantt = ExcelGanttChart()
    
    if include_assets:
        gantt.create_asset_gantt(voyage_data)
    
    if include_berths:
        gantt.create_berth_gantt(voyage_data)
    
    gantt.save(output_file)
