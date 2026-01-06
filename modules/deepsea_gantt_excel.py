"""
Excel Gantt для Deep Sea
========================

Календарное представление рейсов по месяцам
"""

from datetime import datetime, timedelta, date
from typing import List, Dict, Optional
from pathlib import Path
import calendar
import logging

from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

from modules.deepsea_data import DeepSeaData, CalculatedVoyage, CalculatedLeg

logger = logging.getLogger(__name__)


# Цвета по типам операций
COLORS = {
    'loading': '92D050',      # Зелёный
    'discharge': '00B0F0',    # Голубой
    'sea': 'FFC000',          # Оранжевый (laden)
    'sea_ballast': 'D9D9D9',  # Серый (ballast)
    'canal': '7030A0',        # Фиолетовый
    'bunker': '833C0C',       # Коричневый
    'waiting': 'FF0000',      # Красный
}

# Аббревиатуры
ABBREV = {
    'loading': 'L',
    'discharge': 'D',
    'sea': '→',
    'canal': 'C',
    'bunker': 'B',
    'waiting': 'W',
}

FILLS = {k: PatternFill(start_color=v, end_color=v, fill_type='solid') 
         for k, v in COLORS.items()}

HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=9)
WEEKEND_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)


class DeepSeaGanttExcel:
    """Генератор Excel Gantt для Deep Sea"""
    
    def __init__(self, data: DeepSeaData, output_dir: str = "output/deepsea"):
        self.data = data
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_schedule(self, scenario_name: str = "") -> List[str]:
        """Генерация расписания - главная точка входа"""
        logger.info("Генерация Excel Gantt для Deep Sea...")
        
        # Генерируем файлы по месяцам
        month_files = self.generate_all_months(scenario_name)
        
        # Генерируем обзор флота
        fleet_file = self.generate_fleet_overview(scenario_name)
        
        # Возвращаем все пути
        all_files = month_files.copy()
        if fleet_file:
            all_files.append(fleet_file)
        
        return all_files
    
    def generate_all_months(self, scenario_name: str = "") -> List[str]:
        """Генерация Gantt для всех месяцев"""
        if not self.data.calculated_voyages:
            logger.warning("Нет рассчитанных рейсов")
            return []
        
        # Собираем все legs
        all_legs = []
        for voyage in self.data.calculated_voyages.values():
            all_legs.extend(voyage.legs)
        
        if not all_legs:
            return []
        
        # Диапазон месяцев
        min_date = min(leg.start_time for leg in all_legs)
        max_date = max(leg.end_time for leg in all_legs)
        
        files = []
        current = date(min_date.year, min_date.month, 1)
        end = date(max_date.year, max_date.month, 1)
        
        while current <= end:
            filepath = self.generate_month(current.year, current.month, scenario_name)
            if filepath:
                files.append(filepath)
            
            if current.month == 12:
                current = date(current.year + 1, 1, 1)
            else:
                current = date(current.year, current.month + 1, 1)
        
        return files
    
    def generate_month(self, year: int, month: int, scenario_name: str = "") -> Optional[str]:
        """Генерация Gantt для одного месяца"""
        logger.info(f"   Gantt: {year}-{month:02d}")
        
        month_start = datetime(year, month, 1)
        days_in_month = calendar.monthrange(year, month)[1]
        month_end = datetime(year, month, days_in_month, 23, 59, 59)
        
        # Собираем legs за месяц
        month_legs = []
        for voyage in self.data.calculated_voyages.values():
            for leg in voyage.legs:
                if leg.start_time <= month_end and leg.end_time >= month_start:
                    month_legs.append((voyage, leg))
        
        if not month_legs:
            return None
        
        # Группируем по судам
        legs_by_vessel: Dict[str, List[tuple]] = {}
        for voyage, leg in month_legs:
            if leg.vessel_id not in legs_by_vessel:
                legs_by_vessel[leg.vessel_id] = []
            legs_by_vessel[leg.vessel_id].append((voyage, leg))
        
        # Создаём Excel
        wb = Workbook()
        ws = wb.active
        ws.title = f"{year}-{month:02d}"
        
        # === ЗАГОЛОВОК ===
        month_name = calendar.month_name[month]
        title = f"DEEP SEA GANTT: {month_name} {year}"
        if scenario_name:
            title += f" [{scenario_name}]"
        
        ws.cell(row=1, column=1, value=title)
        ws.cell(row=1, column=1).font = Font(bold=True, size=14, color='1F4E79')
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(15, days_in_month + 4))
        
        # === ШАПКА ===
        header_row = 3
        
        # Колонки: Судно | Класс | Рейс | Маршрут | Дни месяца...
        headers = ['Судно', 'Класс', 'Рейс', 'Маршрут']
        for i, h in enumerate(headers):
            cell = ws.cell(row=header_row, column=i + 1, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = THIN_BORDER
        
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 12
        
        # Дни месяца
        for day in range(1, days_in_month + 1):
            col = day + len(headers)
            cell = ws.cell(row=header_row, column=col, value=day)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(col)].width = 3.2
            
            # Выходные
            day_date = date(year, month, day)
            if day_date.weekday() >= 5:
                cell.fill = PatternFill(start_color='4A6FA5', end_color='4A6FA5', fill_type='solid')
        
        # Строка с днями недели
        dow_row = header_row + 1
        for i in range(len(headers)):
            ws.cell(row=dow_row, column=i + 1, value="")
        
        for day in range(1, days_in_month + 1):
            col = day + len(headers)
            day_date = date(year, month, day)
            dow = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс'][day_date.weekday()]
            cell = ws.cell(row=dow_row, column=col, value=dow)
            cell.font = Font(size=7, color='666666')
            cell.alignment = Alignment(horizontal='center')
            if day_date.weekday() >= 5:
                cell.fill = WEEKEND_FILL
        
        # === ДАННЫЕ ===
        data_row = header_row + 2
        
        # Сортируем суда по классу
        vessel_ids = sorted(
            legs_by_vessel.keys(),
            key=lambda v: (
                self.data.vessels.get(v).vessel_class if self.data.vessels.get(v) else 'ZZZ',
                v
            )
        )
        
        for vessel_id in vessel_ids:
            vessel = self.data.vessels.get(vessel_id)
            vessel_legs = legs_by_vessel[vessel_id]
            
            # Группируем по рейсам
            legs_by_voyage: Dict[str, List[CalculatedLeg]] = {}
            voyage_info: Dict[str, CalculatedVoyage] = {}
            
            for voyage, leg in vessel_legs:
                if voyage.voyage_id not in legs_by_voyage:
                    legs_by_voyage[voyage.voyage_id] = []
                    voyage_info[voyage.voyage_id] = voyage
                legs_by_voyage[voyage.voyage_id].append(leg)
            
            # Строка для каждого рейса
            for voyage_id in sorted(legs_by_voyage.keys()):
                voyage = voyage_info[voyage_id]
                voyage_legs = legs_by_voyage[voyage_id]
                
                # Судно
                cell = ws.cell(row=data_row, column=1, 
                              value=vessel.vessel_name if vessel else vessel_id)
                cell.font = Font(size=9, bold=True)
                cell.border = THIN_BORDER
                
                # Класс
                cell = ws.cell(row=data_row, column=2,
                              value=vessel.vessel_class if vessel else '')
                cell.font = Font(size=8)
                cell.border = THIN_BORDER
                
                # Рейс
                cell = ws.cell(row=data_row, column=3, value=voyage_id[-8:])
                cell.font = Font(size=8)
                cell.border = THIN_BORDER
                
                # Маршрут
                route = f"{voyage.load_port}→{voyage.disch_port}"
                cell = ws.cell(row=data_row, column=4, value=route)
                cell.font = Font(size=8)
                cell.border = THIN_BORDER
                
                # Дни
                for day in range(1, days_in_month + 1):
                    col = day + len(headers)
                    day_start = datetime(year, month, day, 0, 0, 0)
                    day_end = datetime(year, month, day, 23, 59, 59)
                    
                    cell = ws.cell(row=data_row, column=col)
                    cell.border = THIN_BORDER
                    
                    # Выходные
                    day_date = date(year, month, day)
                    if day_date.weekday() >= 5:
                        cell.fill = WEEKEND_FILL
                    
                    # Ищем операцию
                    for leg in voyage_legs:
                        if leg.start_time <= day_end and leg.end_time >= day_start:
                            # Определяем цвет
                            if leg.leg_type == 'sea':
                                if leg.cargo_state == 'ballast':
                                    fill_key = 'sea_ballast'
                                else:
                                    fill_key = 'sea'
                            else:
                                fill_key = leg.leg_type
                            
                            fill = FILLS.get(fill_key, FILLS['sea'])
                            cell.fill = fill
                            
                            # Аббревиатура
                            abbrev = ABBREV.get(leg.leg_type, '?')
                            if leg.leg_type == 'sea' and leg.cargo_state == 'ballast':
                                abbrev = '⟶'
                            
                            # Проверяем, что ячейка не является частью объединённого диапазона
                            if not isinstance(cell, MergedCell):
                                cell.value = abbrev
                                cell.font = Font(size=7, bold=True,
                                               color='FFFFFF' if fill_key != 'sea_ballast' else '000000')
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                            break
                
                data_row += 1
            
            # Пустая строка между судами
            data_row += 1
        
        # === ЛЕГЕНДА ===
        legend_row = data_row + 2
        ws.cell(row=legend_row, column=1, value="Легенда:").font = Font(bold=True)
        
        legend_items = [
            ('L - Загрузка', 'loading'),
            ('D - Выгрузка', 'discharge'),
            ('→ - Переход (груз)', 'sea'),
            ('⟶ - Балласт', 'sea_ballast'),
            ('C - Канал', 'canal'),
            ('B - Бункер', 'bunker'),
        ]
        
        legend_col = 2
        for label, color_key in legend_items:
            cell = ws.cell(row=legend_row, column=legend_col, value=label)
            cell.fill = FILLS.get(color_key)
            cell.font = Font(size=8, color='FFFFFF' if color_key != 'sea_ballast' else '000000')
            cell.border = THIN_BORDER
            legend_col += 1
        
        # === СТАТИСТИКА ===
        stats_row = legend_row + 2
        ws.cell(row=stats_row, column=1, value="Статистика месяца:").font = Font(bold=True)
        
        # Считаем статистику
        month_voyages = set()
        month_cargo = 0
        month_distance = 0
        
        for voyage, leg in month_legs:
            month_voyages.add(voyage.voyage_id)
            if leg.leg_type == 'loading':
                month_cargo += leg.qty_mt
            month_distance += leg.distance_nm
        
        stats = [
            f"Рейсов: {len(month_voyages)}",
            f"Груз: {month_cargo:,.0f} MT",
            f"Дистанция: {month_distance:,.0f} nm",
        ]
        
        for i, stat in enumerate(stats):
            ws.cell(row=stats_row, column=2 + i, value=stat).font = Font(size=9)
        
        # Сохраняем
        suffix = f"_{scenario_name}" if scenario_name else ""
        filename = f"gantt_deepsea_{year}_{month:02d}{suffix}.xlsx"
        filepath = self.output_dir / filename
        wb.save(filepath)
        
        logger.info(f"      {filepath}")
        return str(filepath)
    
    def generate_fleet_overview(self, scenario_name: str = "") -> str:
        """Генерация обзора флота - все суда на одном листе"""
        logger.info("  Обзор флота...")
        
        if not self.data.calculated_voyages:
            return ""
        
        # Собираем все legs
        all_legs = []
        for voyage in self.data.calculated_voyages.values():
            all_legs.extend([(voyage, leg) for leg in voyage.legs])
        
        if not all_legs:
            return ""
        
        # Диапазон дат
        min_date = min(leg.start_time for _, leg in all_legs)
        max_date = max(leg.end_time for _, leg in all_legs)
        
        # Округляем до начала/конца месяца
        start_date = date(min_date.year, min_date.month, 1)
        if max_date.month == 12:
            end_date = date(max_date.year, 12, 31)
        else:
            end_date = date(max_date.year, max_date.month + 1, 1) - timedelta(days=1)
        
        total_days = (end_date - start_date).days + 1
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Fleet Overview"
        
        # Заголовок
        title = "DEEP SEA FLEET OVERVIEW"
        if scenario_name:
            title += f" [{scenario_name}]"
        
        ws.cell(row=1, column=1, value=title)
        ws.cell(row=1, column=1).font = Font(bold=True, size=14, color='1F4E79')
        
        ws.cell(row=2, column=1, value=f"Период: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}")
        ws.cell(row=2, column=1).font = Font(size=10, italic=True)
        
        # Шапка
        header_row = 4
        ws.cell(row=header_row, column=1, value="Судно").fill = HEADER_FILL
        ws.cell(row=header_row, column=1).font = HEADER_FONT
        ws.column_dimensions['A'].width = 16
        
        # Недели вместо дней (для компактности)
        current = start_date
        col = 2
        week_cols = []
        
        while current <= end_date:
            week_start = current
            week_end = min(current + timedelta(days=6), end_date)
            
            label = f"{week_start.day:02d}.{week_start.month:02d}"
            cell = ws.cell(row=header_row, column=col, value=label)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(col)].width = 6
            
            week_cols.append((col, week_start, week_end))
            
            current = week_end + timedelta(days=1)
            col += 1
        
        # Группируем по судам
        legs_by_vessel: Dict[str, List[tuple]] = {}
        for voyage, leg in all_legs:
            if leg.vessel_id not in legs_by_vessel:
                legs_by_vessel[leg.vessel_id] = []
            legs_by_vessel[leg.vessel_id].append((voyage, leg))
        
        # Данные
        data_row = header_row + 1
        
        for vessel_id in sorted(legs_by_vessel.keys()):
            vessel = self.data.vessels.get(vessel_id)
            vessel_legs = legs_by_vessel[vessel_id]
            
            # Судно
            cell = ws.cell(row=data_row, column=1, 
                          value=vessel.vessel_name if vessel else vessel_id)
            cell.font = Font(size=9, bold=True)
            cell.border = THIN_BORDER
            
            # Недели
            for col, week_start, week_end in week_cols:
                cell = ws.cell(row=data_row, column=col)
                cell.border = THIN_BORDER
                
                week_start_dt = datetime.combine(week_start, datetime.min.time())
                week_end_dt = datetime.combine(week_end, datetime.max.time())
                
                # Ищем операции за неделю
                week_ops = set()
                for voyage, leg in vessel_legs:
                    if leg.start_time <= week_end_dt and leg.end_time >= week_start_dt:
                        week_ops.add(leg.leg_type)
                
                if week_ops:
                    # Приоритет цвета: loading > discharge > sea > canal > bunker
                    if 'loading' in week_ops:
                        cell.fill = FILLS['loading']
                        cell.value = 'L'
                    elif 'discharge' in week_ops:
                        cell.fill = FILLS['discharge']
                        cell.value = 'D'
                    elif 'sea' in week_ops:
                        cell.fill = FILLS['sea']
                        cell.value = '→'
                    elif 'canal' in week_ops:
                        cell.fill = FILLS['canal']
                        cell.value = 'C'
                    else:
                        cell.fill = FILLS.get('bunker', FILLS['sea'])
                        cell.value = 'B'
                    
                    cell.font = Font(size=8, bold=True, color='FFFFFF')
                    cell.alignment = Alignment(horizontal='center')
            
            data_row += 1
        
        # Утилизация
        util_row = data_row + 2
        ws.cell(row=util_row, column=1, value="Утилизация:").font = Font(bold=True)
        
        for vessel_id in sorted(legs_by_vessel.keys()):
            vessel = self.data.vessels.get(vessel_id)
            vessel_legs = legs_by_vessel[vessel_id]
            
            # Считаем дни занятости
            busy_days = set()
            for voyage, leg in vessel_legs:
                current = leg.start_time.date()
                while current <= leg.end_time.date():
                    busy_days.add(current)
                    current += timedelta(days=1)
            
            utilization = len(busy_days) / total_days * 100
            
            ws.cell(row=util_row, column=1, 
                   value=f"{vessel.vessel_name if vessel else vessel_id}: {utilization:.0f}%")
            util_row += 1
        
        # Сохраняем
        suffix = f"_{scenario_name}" if scenario_name else ""
        filename = f"fleet_overview{suffix}.xlsx"
        filepath = self.output_dir / filename
        wb.save(filepath)
        
        logger.info(f"      {filepath}")
        return str(filepath)