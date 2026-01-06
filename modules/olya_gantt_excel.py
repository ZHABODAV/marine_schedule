"""
Excel Gantt по месяцам
======================

Календарное представление с днями в колонках
"""

from datetime import datetime, timedelta, date
from typing import List, Dict, Optional, Tuple, cast
from pathlib import Path
import calendar
import logging

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

from modules.olya_data import OlyaData, CalculatedOperation, CalculatedVoyage

logger = logging.getLogger(__name__)


# Цвета операций
COLORS = {
    'loading': 'C6EFCE',      # Зелёный (загрузка)
    'discharge': 'BDD7EE',    # Голубой (выгрузка)
    'transit': 'FFE699',      # Жёлтый (переход)
    'waiting': 'F8CBAD',      # Оранжевый (ожидание - можно оптимизировать!)
    'bunkering': 'D9D9D9',    # Серый
}

# Заливки
FILLS = {k: PatternFill(start_color=v, end_color=v, fill_type='solid') 
         for k, v in COLORS.items()}

HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=9)
WEEKEND_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)


class OlyaGanttExcel:
    """
    Генератор Excel Gantt по месяцам
    
    Формат:
    - Строки = суда
    - Колонки = дни месяца (1-31)
    - Ячейки = операции (цветом)
    """
    
    def __init__(self, data: OlyaData, output_dir: str = "output/olya"):
        self.data = data
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_schedule(self) -> List[str]:
        """Генерация расписания - главная точка входа"""
        logger.info("Генерация Excel Gantt для Olya...")
        
        # Генерируем файлы по месяцам
        month_files = self.generate_all_months()
        
        # Генерируем сводный файл
        summary_file = self.generate_summary_gantt()
        
        # Возвращаем все пути
        all_files = month_files.copy()
        if summary_file:
            all_files.append(summary_file)
        
        return all_files
    
    def generate_all_months(self) -> List[str]:
        """Генерация Gantt для всех месяцев с данными"""
        if not self.data.calculated_operations:
            logger.warning("Нет рассчитанных операций")
            return []
        
        # Определяем диапазон месяцев
        min_date = min(op.start_time for op in self.data.calculated_operations)
        max_date = max(op.end_time for op in self.data.calculated_operations)
        
        files = []
        current = date(min_date.year, min_date.month, 1)
        end = date(max_date.year, max_date.month, 1)
        
        while current <= end:
            filepath = self.generate_month(current.year, current.month)
            if filepath:
                files.append(filepath)
            
            # Следующий месяц
            if current.month == 12:
                current = date(current.year + 1, 1, 1)
            else:
                current = date(current.year, current.month + 1, 1)
        
        return files
    
    def generate_month(self, year: int, month: int) -> Optional[str]:
        """Генерация Gantt для одного месяца"""
        logger.info(f"\nГенерация Gantt: {year}-{month:02d}")
        
        # Фильтруем операции за месяц
        month_start = datetime(year, month, 1)
        days_in_month = calendar.monthrange(year, month)[1]
        month_end = datetime(year, month, days_in_month, 23, 59, 59)
        
        # Операции которые попадают в этот месяц
        month_ops = [
            op for op in self.data.calculated_operations
            if op.start_time <= month_end and op.end_time >= month_start
        ]
        
        if not month_ops:
            logger.info(f"   Нет операций за {year}-{month:02d}")
            return None
        
        # Группируем по судам
        ops_by_vessel: Dict[str, List[CalculatedOperation]] = {}
        for op in month_ops:
            if op.vessel_id not in ops_by_vessel:
                ops_by_vessel[op.vessel_id] = []
            ops_by_vessel[op.vessel_id].append(op)
        
        # Создаём Excel
        wb = Workbook()
        ws = cast(Worksheet, wb.active)
        ws.title = f"{year}-{month:02d}"
        
        # === ЗАГОЛОВОК ===
        month_name = calendar.month_name[month]
        ws.cell(row=1, column=1, value=f"GANTT: {month_name} {year}")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(10, days_in_month + 2))
        
        # === ШАПКА С ДНЯМИ ===
        header_row = 3
        
        # Колонка "Судно"
        cell = ws.cell(row=header_row, column=1, value="Судно")
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions['A'].width = 18
        
        # Колонка "Рейс"
        cell = ws.cell(row=header_row, column=2, value="Рейс/Груз")
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        ws.column_dimensions['B'].width = 15
        
        # Дни месяца
        for day in range(1, days_in_month + 1):
            col = day + 2
            cell = ws.cell(row=header_row, column=col, value=day)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
            
            # Ширина колонки
            ws.column_dimensions[get_column_letter(col)].width = 3.5
            
            # Выходные дни (сб, вс) - подсветка в заголовке
            day_date = date(year, month, day)
            if day_date.weekday() >= 5:  # Сб=5, Вс=6
                cell.fill = PatternFill(start_color='7B7B7B', end_color='7B7B7B', fill_type='solid')
        
        # === Строка с днями недели ===
        dow_row = header_row + 1
        ws.cell(row=dow_row, column=1, value="")
        ws.cell(row=dow_row, column=2, value="")
        
        for day in range(1, days_in_month + 1):
            col = day + 2
            day_date = date(year, month, day)
            dow = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс'][day_date.weekday()]
            cell = ws.cell(row=dow_row, column=col, value=dow)
            cell.font = Font(size=7, color='666666')
            cell.alignment = Alignment(horizontal='center')
            
            if day_date.weekday() >= 5:
                cell.fill = WEEKEND_FILL
        
        # === ДАННЫЕ ПО СУДАМ ===
        data_row = header_row + 2
        
        # Сортируем суда: сначала баржи, потом морские
        vessel_ids = sorted(ops_by_vessel.keys(), 
                          key=lambda v: (0 if self.data.vessels.get(v, None) and 
                                        self.data.vessels[v].is_barge else 1, v))
        
        for vessel_id in vessel_ids:
            vessel = self.data.get_vessel(vessel_id)
            vessel_ops = ops_by_vessel[vessel_id]
            
            # Группируем операции по рейсам
            voyages_in_month = {}
            for op in vessel_ops:
                if op.voyage_id not in voyages_in_month:
                    voyages_in_month[op.voyage_id] = []
                voyages_in_month[op.voyage_id].append(op)
            
            # Строка для каждого рейса судна
            for voyage_id, voyage_ops in voyages_in_month.items():
                vessel_name = vessel.vessel_name if vessel else vessel_id
                
                # Колонка судна
                cell = ws.cell(row=data_row, column=1, value=vessel_name)
                cell.font = Font(size=9)
                cell.border = THIN_BORDER
                
                # Колонка рейса/груза
                cargo = voyage_ops[0].cargo if voyage_ops else ""
                qty = max((op.qty_mt for op in voyage_ops if op.qty_mt > 0), default=0)
                voyage_info = f"{cargo} {qty/1000:.0f}k" if qty > 0 else voyage_id[-6:]
                
                cell = ws.cell(row=data_row, column=2, value=voyage_info)
                cell.font = Font(size=8)
                cell.border = THIN_BORDER
                
                # Заполняем дни
                for day in range(1, days_in_month + 1):
                    col = day + 2
                    day_start = datetime(year, month, day, 0, 0, 0)
                    day_end = datetime(year, month, day, 23, 59, 59)
                    
                    cell = cast(Cell, ws.cell(row=data_row, column=col))
                    cell.border = THIN_BORDER

                    # Выходные
                    day_date = date(year, month, day)
                    if day_date.weekday() >= 5:
                        cell.fill = WEEKEND_FILL

                    # Ищем операцию в этот день
                    for op in voyage_ops:
                        if op.start_time <= day_end and op.end_time >=  day_start:
                            # Операция попадает на этот день
                            fill = FILLS.get(op.operation, FILLS.get('transit'))
                            if fill:
                                cell.fill = fill

                            # Аббревиатура операции
                            abbrev = {
                                'loading': 'L',
                                'discharge': 'D',
                                'transit': '→',
                                'waiting': 'W',
                                'bunkering': 'B'
                            }.get(op.operation, '?')

                            cell.value = abbrev
                            cell.font = Font(size=7, bold=True)
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            break
                
                data_row += 1
            
            # Пустая строка между судами
            data_row += 1
        
        # === ЛЕГЕНДА ===
        legend_row = data_row + 2
        ws.cell(row=legend_row, column=1, value="Легенда:")
        ws.cell(row=legend_row, column=1).font = Font(bold=True)
        
        legend_items = [
            ('L - Загрузка', 'loading'),
            ('D - Выгрузка', 'discharge'),
            ('→ - Переход', 'transit'),
            ('W - Ожидание ', 'waiting'),
        ]
        
        legend_col = 2
        for label, op_type in legend_items:
            cell = ws.cell(row=legend_row, column=legend_col, value=label)
            op_fill = FILLS.get(op_type)
            if op_fill:
                cell.fill = op_fill
            cell.font = Font(size=9)
            cell.border = THIN_BORDER
            legend_col += 1
        
        # Сохраняем
        filename = f"gantt_{year}_{month:02d}.xlsx"
        filepath = self.output_dir / filename
        wb.save(filepath)
        
        logger.info(f"    Сохранено: {filepath}")
        return str(filepath)
    
    def generate_summary_gantt(self) -> str:
        """Генерация сводного Gantt на все месяцы в одном файле"""
        logger.info("\n Генерация сводного Gantt...")
        
        if not self.data.calculated_operations:
            return ""
        
        wb = Workbook()
        
        # Определяем диапазон
        min_date = min(op.start_time for op in self.data.calculated_operations)
        max_date = max(op.end_time for op in self.data.calculated_operations)
        
        current = date(min_date.year, min_date.month, 1)
        end = date(max_date.year, max_date.month, 1)
        
        first_sheet = True
        while current <= end:
            self._add_month_sheet(wb, current.year, current.month, first_sheet)
            first_sheet = False
            
            if current.month == 12:
                current = date(current.year + 1, 1, 1)
            else:
                current = date(current.year, current.month + 1, 1)
        
        # Удаляем пустой первый лист если есть
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        filepath = self.output_dir / "gantt_all_months.xlsx"
        wb.save(filepath)
        
        logger.info(f"    Сохранено: {filepath}")
        return str(filepath)
    
    def _add_month_sheet(self, wb: Workbook, year: int, month: int, first_sheet: bool):
        """Добавление листа месяца в книгу"""
        month_start = datetime(year, month, 1)
        days_in_month = calendar.monthrange(year, month)[1]
        month_end = datetime(year, month, days_in_month, 23, 59, 59)
        
        month_ops = [
            op for op in self.data.calculated_operations
            if op.start_time <= month_end and op.end_time >= month_start
        ]
        
        if not month_ops:
            return
        
        # Создаём лист
        month_name = calendar.month_abbr[month]
        sheet_name = f"{month_name} {year}"
        
        if first_sheet:
            ws = cast(Worksheet, wb.active)
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(sheet_name)
        
        # Используем ту же логику что и в generate_month
        # ... (копируем логику заполнения)
        
        # Упрощённая версия
        ops_by_vessel: Dict[str, List[CalculatedOperation]] = {}
        for op in month_ops:
            if op.vessel_id not in ops_by_vessel:
                ops_by_vessel[op.vessel_id] = []
            ops_by_vessel[op.vessel_id].append(op)
        
        # Заголовок
        ws.cell(row=1, column=1, value=f"{calendar.month_name[month]} {year}")
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        
        # Шапка
        header_row = 2
        ws.cell(row=header_row, column=1, value="Судно").font = HEADER_FONT
        ws.cell(row=header_row, column=1).fill = HEADER_FILL
        ws.column_dimensions['A'].width = 15
        
        for day in range(1, days_in_month + 1):
            col = day + 1
            cell = ws.cell(row=header_row, column=col, value=day)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(col)].width = 3
        
        # Данные
        data_row = 3
        for vessel_id in sorted(ops_by_vessel.keys()):
            vessel = self.data.get_vessel(vessel_id)
            vessel_name = vessel.vessel_name if vessel else vessel_id
            
            ws.cell(row=data_row, column=1, value=vessel_name).font = Font(size=8)
            
            for day in range(1, days_in_month + 1):
                col = day + 1
                day_start = datetime(year, month, day, 0, 0, 0)
                day_end = datetime(year, month, day, 23, 59, 59)
                
                cell = cast(Cell, ws.cell(row=data_row, column=col))

                for op in ops_by_vessel[vessel_id]:
                    if op.start_time <= day_end and op.end_time >= day_start:
                        fill = FILLS.get(op.operation, FILLS.get('transit'))
                        if fill:
                            cell.fill = fill
                        break
            
            data_row += 1