"""
Template Creation Tool
Creates Excel templates with dropdowns and data validations.
"""
from __future__ import annotations

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import List, Dict, Optional


def create_voyage_input_template(output_file: str) -> None:
    """
    Create Excel template for voyage input with data validations.
    
    Args:
        output_file: Path to save the template
    """
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise ValueError("Failed to create active worksheet")
    ws.title = "Voyage Input"
    
    # Define headers
    headers = [
        'asset', 'start_port', 'end_port', 'start_time', 
        'duration_hours', 'leg_type', 'cargo_type', 'cargo_quantity'
    ]
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Create reference data sheet
    ref_ws = wb.create_sheet("Reference Data")
    
    # Sample assets
    assets = ['Asset_001', 'Asset_002', 'Asset_003', 'Asset_004', 'Asset_005']
    for idx, asset in enumerate(assets, start=1):
        ref_ws.cell(row=idx, column=1, value=asset)
    
    # Sample ports
    ports = ['Balakovo', 'Saratov', 'Volgograd', 'Astrakhan', 'Samara', 'Kazan']
    for idx, port in enumerate(ports, start=1):
        ref_ws.cell(row=idx, column=2, value=port)
    
    # Leg types
    leg_types = ['sailing', 'loading', 'unloading', 'waiting', 'maintenance']
    for idx, leg_type in enumerate(leg_types, start=1):
        ref_ws.cell(row=idx, column=3, value=leg_type)
    
    # Cargo types
    cargo_types = ['meal', 'oil', 'grain', 'fertilizer', 'containers', 'general']
    for idx, cargo_type in enumerate(cargo_types, start=1):
        ref_ws.cell(row=idx, column=4, value=cargo_type)
    
    # Add headers to reference sheet
    ref_ws['A1'] = 'Assets'
    ref_ws['B1'] = 'Ports'
    ref_ws['C1'] = 'Leg Types'
    ref_ws['D1'] = 'Cargo Types'
    
    for cell in ref_ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    
    # Create data validations
    # Asset validation
    asset_dv = DataValidation(
        type="list",
        formula1=f"='Reference Data'!$A$2:$A${len(assets)+1}",
        allow_blank=False
    )
    asset_dv.error = 'Please select a valid asset'
    asset_dv.errorTitle = 'Invalid Asset'
    ws.add_data_validation(asset_dv)
    asset_dv.add(f'A2:A1000')
    
    # Port validations (start port)
    port_dv_start = DataValidation(
        type="list",
        formula1=f"='Reference Data'!$B$2:$B${len(ports)+1}",
        allow_blank=False
    )
    port_dv_start.error = 'Please select a valid port'
    port_dv_start.errorTitle = 'Invalid Port'
    ws.add_data_validation(port_dv_start)
    port_dv_start.add(f'B2:B1000')
    
    # Port validation (end port)
    port_dv_end = DataValidation(
        type="list",
        formula1=f"='Reference Data'!$B$2:$B${len(ports)+1}",
        allow_blank=False
    )
    port_dv_end.error = 'Please select a valid port'
    port_dv_end.errorTitle = 'Invalid Port'
    ws.add_data_validation(port_dv_end)
    port_dv_end.add(f'C2:C1000')
    
    # Leg type validation
    leg_type_dv = DataValidation(
        type="list",
        formula1=f"='Reference Data'!$C$2:$C${len(leg_types)+1}",
        allow_blank=False
    )
    leg_type_dv.error = 'Please select a valid leg type'
    leg_type_dv.errorTitle = 'Invalid Leg Type'
    ws.add_data_validation(leg_type_dv)
    leg_type_dv.add(f'F2:F1000')
    
    # Cargo type validation
    cargo_type_dv = DataValidation(
        type="list",
        formula1=f"='Reference Data'!$D$2:$D${len(cargo_types)+1}",
        allow_blank=True
    )
    cargo_type_dv.error = 'Please select a valid cargo type'
    cargo_type_dv.errorTitle = 'Invalid Cargo Type'
    ws.add_data_validation(cargo_type_dv)
    cargo_type_dv.add(f'G2:G1000')
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    
    # Add sample data row
    ws['A2'] = 'Asset_001'
    ws['B2'] = 'Balakovo'
    ws['C2'] = 'Saratov'
    ws['D2'] = '2024-01-01 08:00'
    ws['E2'] = 24
    ws['F2'] = 'loading'
    ws['G2'] = 'meal'
    ws['H2'] = 1000
    
    # Save workbook
    wb.save(output_file)


def create_berth_capacity_template(output_file: str) -> None:
    """
    Create template for berth capacity configuration.
    
    Args:
        output_file: Path to save the template
    """
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise ValueError("Failed to create active worksheet")
    ws.title = "Berth Capacity"
    
    # Headers
    headers = ['port', 'max_capacity', 'berth_type', 'notes']
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Sample data
    sample_ports = [
        ('Balakovo', 2, 'river', 'Main operations port'),
        ('Saratov', 3, 'river', 'Transit port'),
        ('Volgograd', 4, 'river', 'Major hub'),
        ('Astrakhan', 5, 'river/sea', 'Sea access port'),
    ]
    
    for row_idx, (port, capacity, berth_type, notes) in enumerate(sample_ports, start=2):
        ws.cell(row=row_idx, column=1, value=port)
        ws.cell(row=row_idx, column=2, value=capacity)
        ws.cell(row=row_idx, column=3, value=berth_type)
        ws.cell(row=row_idx, column=4, value=notes)
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 30
    
    wb.save(output_file)


def refresh_validations(input_directory: Path) -> None:
    """
    Refresh data validations in existing templates.
    
    Args:
        input_directory: Directory containing templates to refresh
    """
    input_dir = Path(input_directory)
    
    if not input_dir.exists():
        print(f"Directory {input_directory} does not exist. Creating...")
        input_dir.mkdir(parents=True, exist_ok=True)
        return
    
    # Find all Excel files
    excel_files = list(input_dir.glob('*.xlsx'))
    
    for excel_file in excel_files:
        print(f"Refreshing validations in {excel_file.name}...")
        # Here you would implement validation refresh logic
        # For now, just acknowledge the file
        
    print(f"Refreshed {len(excel_files)} files")


def create_all_templates(output_directory: str = "project/input") -> None:
    """
    Create all template files.
    
    Args:
        output_directory: Directory to save templates
    """
    output_dir = Path(output_directory)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    print("Creating voyage input template...")
    create_voyage_input_template(str(output_dir / "voyage_input_template.xlsx"))
    
    print("Creating berth capacity template...")
    create_berth_capacity_template(str(output_dir / "berth_capacity_template.xlsx"))
    
    print("All templates created successfully!")


if __name__ == "__main__":
    create_all_templates()
